import numpy as np
from gamma_method import *
from peng_robinson_method import * 
from antoine_method import *
from scipy.optimize import minimize, fsolve, brentq
import matplotlib.pyplot as plt
import openpyxl as pyxl
import os
from time import time


class McThieleEngine:
    def __init__(self,
                 nrtl_calc: GammaNRTL,
                 antoine_calc: ModeloAntoine, 
                 pr_calc: ModeloPengRobinson,
                 mixture: Mixture,
                 Cpl: np.ndarray,
                 deltaH: np.ndarray,
                 x1_E: float,
                 x1_F: float,
                 x1_D: float,
                 P: float):
        print("Método de Mc-Thiele Inicializado")
        # Worker especficiso
        self.nrtl_calc = nrtl_calc
        self.antoine_calc = antoine_calc
        self.mixture = mixture
        self.pr_calc = pr_calc

        # Parametros termofisicos 
        # Calor especifico liquido
        self.Cpl = Cpl
        # Calor latente
        self.deltaH = deltaH

        # Composição do fundo (E), alimentação (F) e esgotamento (D)
        self.x1_E = x1_E
        self.x1_F = x1_F
        self.x1_D = x1_D

        # Pressão do sistema [Pa]
        self.P = P

        # --- Parametros calculados
        self.q = None
        self.feed_line = None
        self.Rmin = None
        self.R = None
        self.rectifying_line = None
        self.stripping_line = None
        self.x_intercept = None # x de intercepção das tres retas
        self.y_intercept = None # y de intercepção das tres retas
        self.stage_data = []


    def bolha_T(self,
                   T: float, 
                   x: np.ndarray,
                   FO: bool = False):
        """
        Calcula a composicao do vapor ou o residuo  para o calculo da temperatura de bolha de uma mistura.

        Este método opera de duas maneiras, ambas controladas pelo parâmetro 'FO'
        A função traz a modelagem de equilíbrio líquido-vapor baseada na Lei de Raoult Modificada. Considera a fase de vapor
        saturado como um _Gás Ideal_. A fase vapor da mistura é considerada real, o método implementado para sua resolução é 
        Peng-Robinson, equanto a solução é, também, considerada real e resolvida pelo modelo NRTL.

        Args:
            T (float): A temperatura Absoluta do sistema [K]
            x (np.ndarray): Vetor com as frações molares da fase líquida
            P_sis (float): A pressão do sistema [Pa]
            FO (bool, optional): Trata-se de uma flah para determinar se a função vai ser usada para obter a temperatura de bolha.
                - False (padrão): O método retorna a composição do vapor calcualda.
                - True: O método atua como uma função objetivo (FO) e retorna o erro quadrático `(1 - Σy)²` para ser
                  usado em um otimizador numérico (como scipy.optimize.minimize, scipy.optimize.fsolve) para encontrar a
                  temperatura de bolha.

        Returns:
            np.ndarray | float:
                - False (padrão): np.ndarray contendo as frações molares do vapor em equilíbrio com o líquido
                - True: float, representado o resíduo quadrático da soma das frações molares
        """
        P_sat = self.antoine_calc.get_vapor_pressure(T=T)
        gamma = np.exp(self.nrtl_calc.get_lnGamma(T=T, fraction=x))

        # Faz uma estimativa inicial de y (Gas ideal)
        y_guess = (x * gamma * P_sat) / self.P
        y_guess /= np.sum(y_guess)
        # Cria um estado
        local_state = State(mixture=self.mixture, T=T, P=self.P, z=None, is_vapor=True)
        for _ in range(20):
            # Calcula o estado termodinâmico do gas (Peng-Robinson)
            local_state.z = np.array(y_guess)
            local_state.params = self.pr_calc.get_params(state=local_state)
            local_state.Z = max(self.pr_calc.get_Z(state=local_state))
            phi = self.pr_calc.get_phi(state=local_state)

            y_old = (x * gamma * P_sat) / (self.P * phi)
            y_new = y_old / np.sum(y_old)

            if np.allclose(y_old, y_new, atol=1e-8):
                break

            y_guess = y_new
        
        if FO:
            soma_y = np.sum(y_old)
            err = (1.0 - soma_y)
            return err
        else:
            return y_guess    

    def get_equilibrium_point(self, x: np.ndarray, T0: float) -> float:
        """
        A função obter um ponto especifico na curva de equilíbrio. Esse métoto é importante para obter a linha de retificação
        com o refluxo mínimo e para aplicar o método Mc-Thiele completo (descer as escadinhas).
        Esse método chama duas vezes o self.bolha_T. A primeira vez com o parametro 'FO' com True, indicando que é uma otimização
        para obter a temperatura de bolha da mistura. A segunda vez com 'FO' padrão (False), para obter a composição da fase vapor
        em questão.

        Args:
            x (np.ndarray): Vetor com composição da fase líquida
            T0 (float): Uma estimativa inicial para obter a temperatura de bolha específica do ponto [K]

        Return: 
            y1 (float): a fração molar do componente mais volátil
        """

        # Encontra a temperatura de bolha
        # T = minimize(fun=self.bolha_T, 
        #              x0=T0, 
        #              args=(x, True), 
        #              method='NELDER-MEAD').x[0]
        T = fsolve(func=self.bolha_T,
                   x0=T0,
                   args=(x, True))[0]
        
        bubble_T_root = lambda temp: self.bolha_T(T=temp, x=x, FO=True)

        T_min = 340 #K
        T_max = 380 #K
        T_bolha = brentq(bubble_T_root, a=T_min, b=T_max)
        # Encontra a composição do vapor
        y1 = self.bolha_T(T=T, x=x)[0]
        
        return y1

    def pinch_point_FO(self, x1_P):
        x1_P = x1_P[0]
        x_P = np.array([x1_P, 1 - x1_P])
        T0 = np.array([300.0]) #K
        # Obtem o ponto na curva de equilibrio
        yp = self.get_equilibrium_point(x=x_P, T0=T0)

        # Aqui eh um calculo numerido da derivada da curva de equilibrio, usado para obter o ponto de pinch
        eps = 0.0001
        xp_pos = np.array([x1_P + eps, 1 - (x1_P + eps)])
        yp_pos = self.get_equilibrium_point(x=xp_pos, T0=T0)
        xp_neg = np.array([x1_P - eps, 1 - (x1_P - eps)])
        yp_neg = self.get_equilibrium_point(x=xp_neg, T0=T0)
        # Derivada numerica
        dyp_dxp = (yp_pos - yp_neg) / (2 * eps)

        err = (yp - self.x1_D - dyp_dxp * (x1_P - self.x1_D))
        return err

    def compute_q(self):
        x_F = np.array([self.x1_F, 1 - self.x1_F])
        T0 = np.array([300.0]) #K
        # T_bolha = minimize(fun=self.bolha_T, x0=T0, args=(x_F, True), method='NELDER-MEAD').x[0]
        T_bolha = fsolve(func=self.bolha_T, x0=T0, args=(x_F, True))[0]
        print(T_bolha)

        Cpl_mix = np.sum(x_F * self.Cpl)
        deltaH_mix = np.sum(x_F * self.deltaH)
        T_F = 22.22 + 273.15 # PONTO DE APOIO MUDARA ARARA ARARA RARARA
        q = 1 + Cpl_mix * (T_bolha - T_F) / deltaH_mix
        print(q)
        self.q = q
    
    def compute_Rmin(self):
        x1_P = fsolve(self.pinch_point_FO, 
                x0=np.array([0.6]))[0]
        print(x1_P)
        x_P = np.array([x1_P, 1 - x1_P])

        y1_P = self.get_equilibrium_point(x=x_P, T0=np.array([300.0]))
        m_min = (y1_P - self.x1_D) / (x1_P - self.x1_D)
        Rmin = m_min / (1 - m_min)
        print(Rmin)
        self.Rmin = Rmin
    
    def get_operating_lines(self, alpha: float=1.5):
        """
        
        Args:
            alpha (float): Um valor para obter o número de refluxo, a relação é R / R_min = alpha, seu valor padrão é 1.5

        """
        # Calcula a reta de alimentacao
        self.feed_line = {
            'slope': self.q / (self.q - 1),
            'intercept': self.x1_F / (1 - self.q),
        }
        self.R = 1.34 # alpha * self.Rmin
        self.rectifying_line = {
            'slope': self.R / (self.R + 1),
            'intercept': self.x1_D / (self.R + 1)
        }
        self.x_intercept = (self.x1_F / (1 - self.q) - self.x1_D / (self.R + 1)) / (self.R / (self.R + 1) - self.q / (self.q - 1))
        self.y_intercept = self.R * self.x_intercept / (self.R + 1) + self.x1_D / (self.R + 1)
        slope_aux = (self.y_intercept - self.x1_E) / (self.x_intercept - self.x1_E)
        self.stripping_line = {
            'slope': slope_aux,
            'intercept': self.x1_E * (1 - slope_aux)
        }

    def find_x_for_y(self, x1, y1_spec):
        x = np.array([x1, 1 - x1])
        T0 = np.array([300.0])
        y1 = self.get_equilibrium_point(x=x, T0=T0)
        err = y1 - y1_spec
        return err
    
    def get_stages(self):
        m_D, b_D = self.rectifying_line['slope'],  self.rectifying_line['intercept']
        m_E, b_E = self.stripping_line['slope'],  self.stripping_line['intercept']
        x_inter = self.x_intercept
        x1_E = self.x1_E # residuo
        # Ponto inicial
        current_x = self.x1_D
        current_y = self.x1_D

        stage_list = []
        n_stage = 0
        # first stage
        while current_x > x1_E:
            n_stage += 1
            print(n_stage, current_x, current_y)
            
            # Calcula a linha horizontal do mc-thiele
            # x_eq = fsolve(func=self.test, x0=current_x, args=(current_y))[0]
            x_eq = brentq(self.find_x_for_y,
                          a=0.001,
                          b=0.999,
                          args=(current_y))
            h_i = (current_x, current_y)
            h_f = (x_eq, current_y)
            h_line = (h_i, h_f)

            # Calcula a linha vertical do mc-thiele
            v_i = h_f
            current_x = x_eq
            
            # Avalia se passou do ponto linha de alimentacao, o que muda a reta de operacao (retificação -> esgotamento)
            if current_x > x_inter:
                current_y = m_D * current_x + b_D
            else:
                current_y = m_E * current_x + b_E

            v_f = (current_x, current_y)
            v_line = (v_i, v_f)
            complete_state = (h_line, v_line)
            stage_list.append(complete_state)

        self.stage_data = stage_list

    def graph(self):
        name = 'water_ethanol_equilibri_0.9727099bar'
        wb = pyxl.load_workbook((os.path.dirname(os.path.abspath(__file__)) + f'\\data\\{name}.xlsx'))
        sheet = wb.active
        x_plot = []
        y_plot = []
        T_plot = []
        for l in filter(None, sheet.iter_rows(min_row=2, values_only=True)):
            x_plot.append(l[0])
            y_plot.append(l[1])
            T_plot.append(l[2])



        # --- Reta de alimentacao
        m_F, b_F = self.feed_line['slope'],  self.feed_line['intercept']
        x_F = np.linspace(self.x1_F, self.x_intercept, 10)
        y_F = m_F * x_F + b_F

        # --- Reta de retificação
        m_D, b_D = self.rectifying_line['slope'],  self.rectifying_line['intercept']
        x_D = np.linspace(self.x_intercept, self.x1_D, 10)
        y_D = m_D * x_D + b_D

        # --- Reta de esgotamento
        m_E, b_E = self.stripping_line['slope'],  self.stripping_line['intercept']
        x_E = np.linspace(self.x1_E, self.x_intercept, 10)
        y_E = m_E * x_E + b_E

        fig, ax = plt.subplots(figsize=(6, 6))

        ax.plot([0,1],[0,1], color='k', linewidth=0.75)
        ax.plot(x_plot, y_plot, color='slategray', linewidth=1.25)
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        # plt.plot(np.array([x1_F, x_al_eq]), np.array([x1_F, y_al_eq]))
        ax.plot(x_F, y_F, color='k', linewidth=1.0)
        ax.plot(x_D, y_D, color='k', linewidth=1.0)
        ax.plot(x_E, y_E, color='k', linewidth=1.0)

        for stage in self.stage_data:
            h_line = stage[0]
            v_line = stage[1]

            x_h = [h_line[0][0], h_line[1][0]]
            y_h = [h_line[0][1], h_line[1][1]]

            x_v = [v_line[0][0], v_line[1][0]]
            y_v = [v_line[0][1], v_line[1][1]]

            ax.plot(x_h, y_h, color='darkred', linewidth=1)
            ax.plot(x_v, y_v, color='darkred', linewidth=1)

        ax.set_xlabel(xlabel=r'$x_{ethanol}\;$')
        ax.set_ylabel(ylabel=r'$y_{ethanol}\;$')
        plt.show()


if __name__ == '__main__':
    # ----- TEMPERATURA E PRESSAO
    T = 352.135 # K
    P = 97270.99 # Pa
    x = np.array([0.45, 0.55])

    # ----- PARAMETROS NRTL
    a12 = -0.8009
    a21 = 3.4578
    b12 = 246.2 
    b21 = -586.1
    alpha = 0.3

    a_ij = np.array([
        [0, a12],
        [a21, 0]
    ])

    b_ij = np.array([
        [0, b12],
        [b21, 0]
    ])

    nrtl_calc = GammaNRTL(list_aij=a_ij, list_bij=b_ij, alpha=alpha) # Fixa os parametros do modelo

    # ----- PARAMETROS ANTOINE
    water_ant = np.array([5.40221, 1838.675, -31.73])
    ethanol_ant = np.array([5.24677, 1598.673, -46.424])
    params = [ethanol_ant, water_ant]
    antoine_calc = ModeloAntoine(params=params) # Fixa os parametros do modelo

    # ----- PARAMETROS PENG-ROBINSON
    water_pr = Component(name='H2O', Tc=647.10, Pc=220.55e5, omega=0.345) 
    ethanol_pr = Component(name='C2H6O', Tc=513.90, Pc=61.48e5, omega=0.645)
    kij = 0.0
    k_ij = np.array([[0, kij],[kij,0]])
    mixture = Mixture([ethanol_pr, water_pr], k_ij=k_ij, l_ij=0.0)    
    pr_calc = ModeloPengRobinson()


    # PEGA DADOS JA SALVOS PARA O EQUILIBRIO
    name = 'water_ethanol_equilibri_0.9727099bar'
    wb = pyxl.load_workbook((os.path.dirname(os.path.abspath(__file__)) + f'\\data\\{name}.xlsx'))
    sheet = wb.active
    x_plot = []
    y_plot = []
    T_plot = []
    


    # Tentativa de Mc-Thiele
    Cpl = np.array([111283.3087, 75433.29859])
    deltaH = np.array([42738814.21, 44095448.08])
    x1_F = 0.17
    x_F = np.array([x1_F, 1 - x1_F])
    x1_D = 0.80
    x1_E = 0.01
    mc_thiele_calc = McThieleEngine(nrtl_calc=nrtl_calc,
                                    antoine_calc=antoine_calc,
                                    pr_calc=pr_calc,
                                    mixture=mixture,
                                    Cpl=Cpl,
                                    deltaH=deltaH,
                                    x1_E=x1_E,
                                    x1_F=x1_F,
                                    x1_D=x1_D,
                                    P=P)
    print('--'*60)
    t_0 = time()
    mc_thiele_calc.compute_q()
    mc_thiele_calc.compute_Rmin()
    mc_thiele_calc.get_operating_lines()
    mc_thiele_calc.get_stages()
    mc_thiele_calc.graph()

    t_f = time()
    print(t_f - t_0)